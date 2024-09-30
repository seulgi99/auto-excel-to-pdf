from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import re

def check_line(sheet):
    max_row = sheet.max_row
    max_row = int(max_row)
    print(f"데이터가 존재하는 행의 수: {max_row}")
    return max_row


# 거래내역 조회서 자동화에서만 쓰는 것임
def check_max_index(sheet, column, start_row, end_row):
    max_index = None

    # 데이터 탐색 및 최대값 찾기
    for row in range(start_row, end_row + 1):
        cell_value = sheet[f'{column}{row}'].value
        if cell_value is not None:
            if max_index is None or cell_value > max_index:
                max_index = cell_value
        else:
            raise Exception(f'{column}{row}이 비었거나 숫자가 아닙니다.')

    print(f'최종 분할되어 나올 pdf의 개수: {max_index}')
    return max_index


# 거래내역 조회서 자동화에서만 쓰는 것임
def get_target_row_list(sheet, index, column, start_row, end_row):
    target_row_list = []

    # 데이터 탐색 및 해당 인덱스가 들어가있는 열 찾기
    for row in range(start_row, end_row + 1):
        cell_value = sheet[f'{column}{row}'].value
        cell_value = int(cell_value)
        if cell_value == index:
            target_row_list.append(row)
    if not target_row_list:
        raise Exception(f'{index}번이 들어있지 않습니다.')
    return target_row_list


def color_cell_yellow(sheet, row, column):
    cell = sheet[column + str(row)]
    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 노란색 배경
    # 셀에 배경 색상 적용
    cell.fill = fill


def delete_column(sheet, column, start_row, end_row):
    for row in range(start_row, end_row + 1):
        cell = column + str(row)
        sheet[cell].value = None


def get_filtered_data(filename, sheet_name, target_column, filter_value):
    # 엑셀 파일 열기
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]

    # 필터링된 행들을 저장할 리스트
    filtered_rows = []

    # 헤더 포함 행 번호 시작 (일반적으로 1행은 헤더)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        # Q 열의 값 확인 (Q 열은 17번째 열에 해당)
        if row[target_column - 1].value == filter_value:
            # 조건에 맞는 행을 필터링된 리스트에 추가
            filtered_rows.append([cell.value for cell in row])

    workbook.close()

    return filtered_rows


# 입금, 미수금 입금 데이터 가져오기
def get_history_data_set_info(data_set):
    # 엑셀 파일 열기
    workbook = load_workbook(data_set)
    sheet = workbook.active  # 활성 시트 선택

    # 입금 데이터 저장을 위한 리스트
    deposit_data = []
    # 미수금 입금 데이터 저장을 위한 리스트
    received_data = []

    # 모든 행을 반복하면서 A부터 AG열의 값을 가져오기
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=33):  # A부터 AG까지 33개의 열
        # Q 열의 값 확인 (Q 열은 17번째 열에 해당)
        if row[16].value == '대여료및사용료' and '미수금 설정' not in row[4].value:  # Q 열은 인덱스 16에 해당 - 대여료 및 사용료인 것들만 뽑아오기, 미수금 설정 건은 별도로 처리
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            deposit_data.append(row_data)
        if '미수금 입금' in row[4].value and row[8].value == 0:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            received_data.append(row_data)

    workbook.close()

    return deposit_data, received_data


def get_file_info(file):
    # 엑셀 파일 열기
    workbook = load_workbook(file)
    sheet = workbook.active  # 활성 시트 선택

    # 업체 정보를 받기 위한 리스트
    file_data = []

    # 빈 행이 나올 때 까지 모든 행을 반복하면서 B부터 F열의 값을 가져오기
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=6):  # B부터 F까지 5개의 열
        if row[0].value == None:
            break
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        file_data.append(row_data)

    workbook.close()

    return file_data


def get_column_data(filename, column):
    # 엑셀 파일 열기
    workbook = load_workbook(filename)
    sheet = workbook.active  # 활성 시트 선택

    # 데이터 저장을 위한 리스트
    data = []

    # 특정 열의 데이터 가져오기
    for row in range(2, 69):  # 2행부터 68행까지
        cell_value = sheet[f'{column}{row}'].value
        data.append(cell_value)

    workbook.close()

    return data


def get_column_from_date(sheet, date_text):
    # 날짜 형식에서 연도와 월 추출
    match = re.match(r'(\d{4})-(\d{1,2})', date_text)
    if not match:
        raise ValueError("날짜 형식이 잘못되었습니다. 형식을 다음과 같이 맞춰주세요 : YYYY-MM")

    year = match.group(1)[2:]  # 연도의 마지막 두 자리
    month = str(int(match.group(2))) + '월'  # 앞에 '0'을 제거한 월

    # G열부터 탐색
    start_column = 6  # G열의 인덱스는 6
    for col in range(start_column, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value  # 첫 행에서 찾기
        if cell_value and f"{year}.{month}" in cell_value:
            return col

    raise ValueError(f"No column found for date {date_text}")
