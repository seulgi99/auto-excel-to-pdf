import openpyxl
from openpyxl import Workbook
from function import func_excel, func_pdf

def execute(file, data_set) :
    # data set 가져오기
    data_set_info = get_data_set_info(data_set)

    workbook = openpyxl.load_workbook(file)
    source_sheet = workbook.active

    end_row = func_excel.check_line(source_sheet) #데이터가 존재하는 마지막 열

    # I열에서 가장 큰수 찾기
    try:
        max_index = func_excel.check_max_index(source_sheet,'J', 8, end_row)
    except Exception as e:
        workbook.close()
        print(e)
        return

    workbook.close()

    for i in range(1, max_index + 1):
        try:
            print(f'{i}번째 파일 작업')
            # 원본 엑셀 파일 열기
            workbook = openpyxl.load_workbook(file)
            source_sheet = workbook.active

            # 색칠해야하는 열 타게팅
            target_row = func_excel.get_target_row_list(source_sheet, i, 'J', 8, end_row)

            # 새로운 엑셀에 색칠하기(노란색)
            for row in target_row:
                func_excel.color_cell_yellow(source_sheet, row, 'A')
                func_excel.color_cell_yellow(source_sheet, row, 'E')
                func_excel.color_cell_yellow(source_sheet, row, 'F')

            # 이름 지정
            name_cell = 'K' + str(target_row[0])
            # 월 수 지정
            month_cell = 'L' + str(target_row[0])
            month_str = str(source_sheet[month_cell].value) + "월 "
            if source_sheet[month_cell].value == "" or source_sheet[month_cell].value is None:
                month_str = ""
            # 구분(입금/출금) 지정
            type_cell = 'M' + str(target_row[0])

            # 이름 + 월 + 구분으로 파일 이름 지정
            file_name = source_sheet[name_cell].value + " " + month_str + source_sheet[type_cell].value

            # 금액모아서 data set 과 비교
            total_money = 0
            for row in target_row:
                total_money += int(source_sheet["E" + str(row)].value)

            no_data = True # 데이터가 없는 지
            not_match = True # 금액이 일치하는 지
            for data in data_set_info:
                if source_sheet[name_cell].value == data[2]:
                    no_data = False
                    if total_money == int(data[1]):
                        not_match = False
                        break
            if no_data:
                file_name +=" - 데이터 없음"
            elif not_match:
                file_name +=" - 불일치"

            new_excel_name = './excel/' + file_name + '.xlsx'
            new_pdf_name = './pdf/' + file_name + '.pdf'
            # 필요없는 데이터 지우기 ( index 및 파일명,월,구분)
            func_excel.delete_column(source_sheet, 'J', 8, end_row)
            func_excel.delete_column(source_sheet, 'K', 8, end_row)
            func_excel.delete_column(source_sheet, 'L', 8, end_row)
            func_excel.delete_column(source_sheet, 'M', 8, end_row)


            # 새로운 엑셀 파일 저장
            workbook.save(new_excel_name)
            print(new_excel_name + '파일 생성 완료')
            # 엑셀 파일 닫기
            workbook.close()

            func_pdf.excel_to_pdf(new_excel_name, new_pdf_name)
        except Exception as e:
            workbook.close()
            print(f'{i}번째 파일에서 문제가 발생했습니다.')
            print(e)
            return

    func_pdf.merge_pdf()
    input('작업이 완료되었습니다. 엔터를 누르면 메뉴로 돌아갑니다.')


def get_data_set_info(data_set):
    # data set 가져오기
    workbook = openpyxl.load_workbook(data_set)

    data_set_sheet = workbook["일반 데이터"]

    end_row = func_excel.check_line(data_set_sheet)  # 데이터가 존재하는 마지막 열
    data_set_info = []
    for i in range(end_row - 2):
        line = []
        row = i + 3
        line.append(str(data_set_sheet[f'A{row}'].value)) # 상호명
        line.append(str(data_set_sheet[f'B{row}'].value)) # 금액
        line.append(str(data_set_sheet[f'C{row}'].value)) # 생성할 파일 명
        line.append(str(data_set_sheet[f'D{row}'].value)) # 월
        data_set_info.append(line)

    workbook.close()
    return data_set_info