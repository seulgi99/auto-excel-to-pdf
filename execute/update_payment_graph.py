from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from function import func_excel, func_pdf

import re

# 결의서 제목에서 월자 추출 제대로 되는지 확인하기
# 씨앤코컴퍼니, 유니룩스, 스튜디오홍대 창고, 이경선, 윤영(예금), 박기철, 김정현, 허윤석, 김이중(호수 불일치), 송규만(호수 불일치)
# 유니룩스, 숨코리아 빼고 다 없음
# 미수금 반영?: 미수금 설정은 반영x
# 여러 연도 (24.12 ~ 25.1) 일 경우 체크하기
# 납부현황, 원장 관리코드명 통일
# 연구실 동호수
# 파브리카 정산 예외

# file : 그래프를 그리는 파일, data_set : 정보를 가져오는 파일
def execute(file, data_set):
    # deposit_data: 계정과목명이 대여료및사용료 인 것만 중 미수금 설정 제외하고 가져오기 (엑셀 파일 형식의 2차원배열)
    # received_data: 미수금 입금 데이터 가져오기
    deposit_data, received_data = func_excel.get_history_data_set_info(data_set)

    # 그래프의 업체 정보 가져오기
    file_data = func_excel.get_file_info(file)

    # 업체 명, 금액 리스트 표시
    # for keyword, money in file_data[:,0], file_data[:,3]:
    #     print(keyword+': '+money)

    # tenant 임차인
    for idx, tenant in enumerate(file_data):
        name = tenant[0]
        money = str(tenant[3])
        category = tenant[1]
        print(name + ' ' + tenant[1] + ' ' + tenant[2] + '(' + money + '원) >>>')

        # 일반 입금 처리
        for resolution in deposit_data:
            res_code = resolution[19]  # dataset의 관리코드명 (T열) - 그래프 파일의 검색어와 동일
            res_money_value = resolution[8]  # dataset의 대변 (I열)
            res_title = resolution[4]  # dataset의 결의서 제목 (E열)
            res_date = resolution[2]  # dataset의 전표일자 (C열)

            # 일반 업체
            if name == res_code and money == str(res_money_value):
                row = idx + 2
                print("전표일자: ", res_date)
                insert_result = format_date(res_date)
                print("insert result : ", insert_result)  # insert_result : n.n입
                insert_value_to_merged_cell(file, row, res_title, res_date, insert_result)
                print('/////')

            # 산학협력
            if category == "산학협력단":
                room = name.split()
                if room[2] in res_title and room[3] in res_title and money == str(res_money_value):
                    row = idx + 2
                    print("전표일자: ", res_date)
                    insert_result = format_date(res_date)
                    print("insert result : ", insert_result)  # insert_result : n.n입
                    insert_value_to_merged_cell(file, row, res_title, res_date, insert_result)
                    print('/////')

        # 미수금 입금 처리
        for resolution in received_data:
            res_code = resolution[19]  # dataset의 관리코드명 (T열) - 그래프 파일의 검색어와 동일
            res_money_value = resolution[7]  # dataset의 차변 (H열)
            res_title = resolution[4]  # dataset의 결의서 제목 (E열)
            res_date = resolution[2]  # dataset의 전표일자 (C열)
            if name == res_code and money == res_money_value:
                row = idx + 2
                print("**미수금 입금**")
                print("전표일자: ", res_date)
                insert_result = format_date(res_date)
                print("insert result : ", insert_result)  # insert_result : n.n입
                insert_value_to_merged_cell(file, row, res_title, res_date, insert_result)

def extract_and_compute_difference(text):
    # 정규 표현식 패턴
    patterns = [
        r'(\d{2}.\d{1,2}~\d{2}.\d{1,2}월분)', # yy.n~yy.m월분 형태
        r'(\d{1,2}~\d{1,2}월분)',  # n~m월분 형태
        r'(\d{1,2}월분)'  # n월분 형태
    ]

    numbers = []

    for idx, pattern in enumerate(patterns):
        matches = re.findall(pattern, text) # 정규 표현식과 맞는 문자열 리스트로 반환
        if matches:
            # yy.n ~ yy.m, 숫자들 분리 후 2,4 번째 숫자만 추출
            if idx==0:
                annual_parts = re.findall(r'\d{1,2}', matches[0])
                numbers.extend(map(int, annual_parts))
                del numbers[0]  # [1, 2, 3, 4] -> [2, 3, 4]
                del numbers[1]  # [2, 3, 4] -> [2, 4]
                numbers[1] = numbers[1]+12  # 월수 차이 계산 위해 끝나는 월에 +13
                print(pattern)
                print(text)
                print('년')
                print(*numbers)
                break
            # n~m월분 형태인 경우, 시작과 끝 숫자를 분리
            elif idx==1:
                range_parts = re.findall(r'\d+', matches[0])
                numbers.extend(map(int, range_parts))
                print(pattern)
                print(text)
                print('분기')
                print(*numbers)
                break
            # # n월분 형태인 경우, 숫자만 추출
            elif idx==2:
                month_parts = re.findall(r'\d+', matches[0])
                numbers.extend(map(int, month_parts))
                print(pattern)
                print(text)
                print('월')
                print(*numbers)
                break

    if len(numbers) == 0:
        print('결의서 제목의 날짜 양식이 맞지 않습니다.')

    # 숫자가 여러 개일 경우, 차이를 계산
    if len(numbers) > 1:
        difference = max(numbers) - min(numbers)
    else:
        difference = 0  # 숫자가 하나거나 없으면 차이는 0으로 설정

    return min(numbers), difference

# insert_result: n.n입, res_date: dataset의 전표일자, text: dataset의 결의서제목
def insert_value_to_merged_cell(filename, row, text, res_date, insert_result):
    # 엑셀 파일 열기
    workbook = openpyxl.load_workbook(filename, read_only= False)
    sheet = workbook["서울캠"]  # 활성 시트 선택

    # extract_and_compute_difference 함수 호출
    date, difference = extract_and_compute_difference(text)

    # 월수 차이에 따라 병합할 셀 수 설정
    cells_to_merge = difference+1

    print("cells_to_merge", cells_to_merge)

    # 날짜에 맞는 열 찾기
    # date_str = res_date.strftime('%Y-%m-%d')
    date_str = '2024-' + str(date)   # yyyy-mm
    print("date_str : ", date_str)

    column = func_excel.get_column_from_date(sheet, date_str)

    print("cell : ", cell_address(row, column))

    cell = sheet.cell(row=row, column=column)

    print("insert_result : ", insert_result)

    # 지정된 행(row)과 찾은 열(column)에 값 입력
    if cell.value is None:  # 셀이 비어있다면

        print("cell row = " + str(row))
        print("cell column = " + str(column))

        # 필요한 칸 수만큼 셀 병합
        if cells_to_merge > 1:
            sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + cells_to_merge - 1)

        sheet.cell(row=row, column=column).value = insert_result

        # 변경 내용을 저장
        workbook.save(filename)

    # 파일 닫기
    workbook.close()


def cell_address(row, column):
    # 열 번호를 열 문자로 변환
    column_letter = openpyxl.utils.get_column_letter(column)
    # 셀 주소 반환
    return f"{column_letter}{row}"


def format_date(date_str):
    print("date_str : ", date_str)
    # 문자열을 datetime 객체로 변환

    # 연, 월, 일을 각각 추출
    year = date_str.year
    month = date_str.strftime('%m')
    day = date_str.strftime('%d')

    # 변환된 문자열을 반환
    formatted_date = f"{month}.{day}입"
    return formatted_date
